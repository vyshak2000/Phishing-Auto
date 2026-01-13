import extract_msg
import hashlib
import re
import requests
import base64
import os
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
import subprocess
import platform


# ---------------------------------------------------------
# LOAD API KEYS
# ---------------------------------------------------------
def load_api_keys(file_path="api_keys.txt"):
    api_keys = {}
    with open(file_path, "r") as f:
        for line in f:
            if "=" in line:
                key, value = line.strip().split("=", 1)
                api_keys[key] = value
    return api_keys


# ---------------------------------------------------------
# PARSE .MSG FILE
# ---------------------------------------------------------
def parse_msg_file(file_path):
    msg = extract_msg.Message(file_path)

    headers = {
        "From": msg.sender,
        "To": msg.to,
        "Subject": msg.subject,
        "Date": msg.date,
        "Raw-Headers": msg.header if msg.header else ""
    }

    body = msg.body

    # Extract URLs
    urls = re.findall(r"(https?://[^\s]+)", body)

    # Extract attachments
    attachments = []
    for att in msg.attachments:
        content = att.data
        sha256_hash = hashlib.sha256(content).hexdigest()
        attachments.append({
            "filename": att.longFilename,
            "sha256": sha256_hash,
            "size": len(content),
            "content": content
        })

    return headers, body, urls, attachments


# ---------------------------------------------------------
# VIRUSTOTAL URL LOOKUP
# ---------------------------------------------------------
def vt_url_lookup(api_key, url):
    try:
        encoded = base64.urlsafe_b64encode(url.encode()).decode().strip("=")
        r = requests.get(
            f"https://www.virustotal.com/api/v3/urls/{encoded}",
            headers={"x-apikey": api_key}
        )
        if r.status_code != 200:
            return {"error": f"VT URL Error {r.status_code}"}
        return r.json()["data"]["attributes"]["last_analysis_stats"]
    except Exception as e:
        return {"error": str(e)}


# ---------------------------------------------------------
# VIRUSTOTAL HASH LOOKUP
# ---------------------------------------------------------
def vt_hash_lookup(api_key, sha256):
    try:
        r = requests.get(
            f"https://www.virustotal.com/api/v3/files/{sha256}",
            headers={"x-apikey": api_key}
        )
        if r.status_code != 200:
            return {"error": f"VT Hash Error {r.status_code}"}
        return r.json()["data"]["attributes"]["last_analysis_stats"]
    except Exception as e:
        return {"error": str(e)}


# ---------------------------------------------------------
# OTX LOOKUP
# ---------------------------------------------------------
def otx_lookup(api_key, indicator):
    try:
        for t in ["IPv4", "url", "file"]:
            url = f"https://otx.alienvault.com/api/v1/indicators/{t}/{indicator}/general"
            r = requests.get(url, headers={"X-OTX-API-KEY": api_key})
            if r.status_code == 200:
                return r.json()
        return {"error": "OTX Not Found"}
    except Exception as e:
        return {"error": str(e)}


# ---------------------------------------------------------
# ABUSEIPDB LOOKUP
# ---------------------------------------------------------
def abuseipdb_lookup(api_key, ip):
    try:
        r = requests.get(
            "https://api.abuseipdb.com/api/v2/check",
            headers={"Key": api_key, "Accept": "application/json"},
            params={"ipAddress": ip, "maxAgeInDays": 30}
        )
        if r.status_code != 200:
            return {"error": f"AbuseIPDB Error {r.status_code}"}
        return r.json()["data"]
    except Exception as e:
        return {"error": str(e)}


# ---------------------------------------------------------
# BEAUTIFY HELPERS
# ---------------------------------------------------------
def beautify_vt(data):
    if "error" in data:
        return "Error"
    return f"Malicious:{data.get('malicious',0)} Suspicious:{data.get('suspicious',0)} Harmless:{data.get('harmless',0)}"


def beautify_otx(data):
    if "error" in data:
        return "Error"
    pulses = data.get("pulse_info", {}).get("count", 0)
    return f"Pulses:{pulses} Sections:{','.join(data.get('sections',[]))}"


def beautify_abuseip(data):
    if "error" in data:
        return "Error"
    return (
        f"Score:{data.get('abuseConfidenceScore',0)} "
        f"Public:{data.get('isPublic',False)} "
        f"Country:{data.get('countryCode','N/A')}"
    )


# ---------------------------------------------------------
# COLOR CODING
# ---------------------------------------------------------
def apply_color(cell, vt_text=None, abuse_text=None):
    red = PatternFill(start_color="FFC7CE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", fill_type="solid")
    green = PatternFill(start_color="C6EFCE", fill_type="solid")

    if vt_text and "Error" not in vt_text:
        mal = int(vt_text.split("Malicious:")[1].split()[0])
        sus = int(vt_text.split("Suspicious:")[1].split()[0])
        if mal > 0:
            cell.fill = red
        elif sus > 0:
            cell.fill = yellow
        else:
            cell.fill = green

    if abuse_text and "Error" not in abuse_text:
        score = int(abuse_text.split("Score:")[1].split()[0])
        if score > 50:
            cell.fill = red
        elif score > 0:
            cell.fill = yellow
        else:
            cell.fill = green


# ---------------------------------------------------------
# PARSE RAW HEADERS INTO KEY FIELDS
# ---------------------------------------------------------
def extract_field(raw, key):
    for line in raw.split("\n"):
        if line.lower().startswith(key.lower()):
            return line.split(":", 1)[1].strip()
    return "Not Found"


def parse_raw_headers(raw):
    parsed = {}

    parsed["SPF"] = extract_field(raw, "Received-SPF")
    parsed["Authentication-Results"] = extract_field(raw, "Authentication-Results")
    parsed["DKIM"] = extract_field(raw, "DKIM-Signature")
    parsed["DMARC"] = extract_field(raw, "DMARC")
    parsed["Return-Path"] = extract_field(raw, "Return-Path")
    parsed["Content-Type"] = extract_field(raw, "Content-Type")
    parsed["Message-ID"] = extract_field(raw, "Message-ID")
    parsed["X-Trellix"] = extract_field(raw, "X-Trellix")
    parsed["X-IronPort"] = extract_field(raw, "X-IronPort-Anti-Spam-Filtered")
    parsed["X-ThreatScanner"] = extract_field(raw, "X-ThreatScanner-Verdict")

    # Extract sender IP
    ips = re.findall(r"\b(?:[0-9]{1,3}\.){3}[0-9]{1,3}\b", raw)
    parsed["Sender-IP"] = ips[-1] if ips else "Not Found"

    return parsed


# ---------------------------------------------------------
# MULTITHREADED REPUTATION REQUESTS
# ---------------------------------------------------------
def parallel_execute(func, items, *args):
    results = []
    with ThreadPoolExecutor(max_workers=5) as executor:
        mapper = {executor.submit(func, *args, item): item for item in items}
        for future in as_completed(mapper):
            item = mapper[future]
            try:
                results.append((item, future.result()))
            except Exception as e:
                results.append((item, {"error": str(e)}))
    return results


# ---------------------------------------------------------
# CREATE EXCEL REPORT (3 sheets)
# ---------------------------------------------------------
def write_to_excel(headers, parsed_fields, url_results, ip_result, attachment_results, output="email_report.xlsx"):
    wb = Workbook()

    # ---------------- SHEET 1: RAW HEADERS ----------------
    ws1 = wb.active
    ws1.title = "Raw Headers"
    ws1["A1"] = "Raw Headers"
    ws1["A2"] = headers["Raw-Headers"]
    ws1["A2"].alignment = Alignment(wrapText=True)
    ws1.merge_cells("A2:D200")

    # ---------------- SHEET 2: PARSED HEADER ANALYSIS ----------------
    ws2 = wb.create_sheet("Header Analysis")
    ws2.append(["Field", "Value"])

    for k, v in parsed_fields.items():
        ws2.append([k, v])

    # ---------------- SHEET 3: REPUTATION ----------------
    ws3 = wb.create_sheet("Reputation")
    ws3.append(["Type", "Indicator", "VirusTotal", "AbuseIPDB", "OTX"])

    row = 2

    # URL Reputation
    for url, vt, otx in url_results:
        vt_text = beautify_vt(vt)
        otx_text = beautify_otx(otx)
        ws3.append(["URL", url, vt_text, "N/A", otx_text])
        apply_color(ws3[f"C{row}"], vt_text)
        row += 1

    # IP Reputation
    if ip_result:
        ip, vt_ip, abuse, otx_ip = ip_result
        vt_text = beautify_vt(vt_ip)
        abuse_text = beautify_abuseip(abuse)
        otx_text = beautify_otx(otx_ip)

        ws3.append(["IP", ip, vt_text, abuse_text, otx_text])
        apply_color(ws3[f"C{row}"], vt_text)
        apply_color(ws3[f"D{row}"], None, abuse_text)
        row += 1

    # Attachment Hash Reputation
    for sha, vt, otx in attachment_results:
        vt_text = beautify_vt(vt)
        otx_text = beautify_otx(otx)
        ws3.append(["Attachment", sha, vt_text, "N/A", otx_text])
        apply_color(ws3[f"C{row}"], vt_text)
        row += 1

    # ---------- FORMATTING: Borders + Header Style ----------
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color="4B0082", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for sheet in [ws2, ws3]:
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border

        for r in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column):
            for cell in r:
                cell.border = border

    # Auto column width
    for sheet in [ws2, ws3]:
        for col in sheet.columns:
            max_len = max(len(str(c.value)) if c.value else 0 for c in col)
            sheet.column_dimensions[col[0].column_letter].width = max_len + 3

    wb.save(output)
    print(f"[+] Excel report generated: {output}")

    # Auto-open file
    try:
        if platform.system() == "Windows":
            os.startfile(output)
        elif platform.system() == "Darwin":
            subprocess.call(["open", output])
        else:
            subprocess.call(["xdg-open", output])
    except:
        print("[!] Could not open Excel automatically.")


# ---------------------------------------------------------
# MAIN EXECUTION LOGIC
# ---------------------------------------------------------
if __name__ == "__main__":

    input_file = "sample.msg"  # MODIFY HERE

    api = load_api_keys()

    headers, body, urls, attachments = parse_msg_file(input_file)

    # -------- Parse Raw Headers Into Key Fields --------
    raw_header_text = headers.get("Raw-Headers", "")
    parsed_fields = parse_raw_headers(raw_header_text)

    # -------- URL Reputation --------
    vt_url_res = parallel_execute(vt_url_lookup, urls, api["VIRUSTOTAL_API_KEY"])
    otx_url_res = parallel_execute(otx_lookup, urls, api["OTX_API_KEY"])
    url_results = [(u, vt, otx) for (u, vt), (_, otx) in zip(vt_url_res, otx_url_res)]

    # -------- IP Reputation --------
    sender_ip = parsed_fields["Sender-IP"]
    ip_result = None
    if sender_ip != "Not Found":
        abuse = abuseipdb_lookup(api["ABUSEIPDB_API_KEY"], sender_ip)
        otx_ip = otx_lookup(api["OTX_API_KEY"], sender_ip)
        vt_ip = {"error": "VT does not support IP reputation"}
        ip_result = (sender_ip, vt_ip, abuse, otx_ip)

    # -------- Attachment Reputation --------
    hashes = [a["sha256"] for a in attachments]
    vt_att = parallel_execute(vt_hash_lookup, hashes, api["VIRUSTOTAL_API_KEY"])
    otx_att = parallel_execute(otx_lookup, hashes, api["OTX_API_KEY"])
    attachment_results = [(sha, vt, otx) for (sha, vt), (_, otx) in zip(vt_att, otx_att)]

    # -------- Write Excel File --------
    write_to_excel(headers, parsed_fields, url_results, ip_result, attachment_results)
