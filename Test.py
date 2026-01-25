import extract_msg
import re
import hashlib
import requests
import base64
import os
import platform
import subprocess
import ipaddress

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Font

# ==================================================
# LOAD API KEYS
# ==================================================

def load_api_keys(file_path="api_keys.txt"):
    keys = {}
    with open(file_path, "r", encoding="utf-8") as f:
        for line in f:
            if "=" in line:
                k, v = line.strip().split("=", 1)
                keys[k] = v
    return keys

# ==================================================
# EXTRACT URLs & ATTACHMENT HASHES FROM MSG
# ==================================================

def extract_msg_indicators(msg_path):
    msg = extract_msg.Message(msg_path)
    body = msg.body or ""

    urls = re.findall(r'https?://[^\s<>"]+', body)

    attachments = []
    for att in msg.attachments:
        sha = hashlib.sha256(att.data).hexdigest()
        attachments.append(sha)

    return list(set(urls)), list(set(attachments))

# ==================================================
# EXTRACT PUBLIC IPs FROM RAW HEADERS
# ==================================================

def extract_public_ips(headers_path):
    with open(headers_path, "r", encoding="utf-8", errors="ignore") as f:
        data = f.read()

    ips = re.findall(r'\b(?:\d{1,3}\.){3}\d{1,3}\b', data)

    public_ips = []
    for ip in ips:
        try:
            ip_obj = ipaddress.ip_address(ip)
            if ip_obj.is_global:
                public_ips.append(ip)
        except ValueError:
            pass

    return list(set(public_ips))

# ==================================================
# VIRUSTOTAL LOOKUPS
# ==================================================

def vt_url(api_key, url):
    try:
        encoded = base64.urlsafe_b64encode(url.encode()).decode().strip("=")
        r = requests.get(
            f"https://www.virustotal.com/api/v3/urls/{encoded}",
            headers={"x-apikey": api_key},
            timeout=15
        )

        if r.status_code != 200:
            return f"VT URL Error {r.status_code}"

        stats = r.json()["data"]["attributes"]["last_analysis_stats"]
        return f"M:{stats['malicious']} S:{stats['suspicious']} H:{stats['harmless']}"

    except Exception as e:
        return f"Error: {e}"

def vt_hash(api_key, sha):
    try:
        r = requests.get(
            f"https://www.virustotal.com/api/v3/files/{sha}",
            headers={"x-apikey": api_key},
            timeout=15
        )

        if r.status_code != 200:
            return f"VT Hash Error {r.status_code}"

        stats = r.json()["data"]["attributes"]["last_analysis_stats"]
        return f"M:{stats['malicious']} S:{stats['suspicious']} H:{stats['harmless']}"

    except Exception as e:
        return f"Error: {e}"

def vt_ip(api_key, ip):
    try:
        r = requests.get(
            f"https://www.virustotal.com/api/v3/ip_addresses/{ip}",
            headers={"x-apikey": api_key},
            timeout=15
        )

        if r.status_code != 200:
            return f"VT IP Error {r.status_code}"

        stats = r.json()["data"]["attributes"]["last_analysis_stats"]
        return f"M:{stats['malicious']} S:{stats['suspicious']} H:{stats['harmless']}"

    except Exception as e:
        return f"Error: {e}"

# ==================================================
# EXCEL OUTPUT
# ==================================================

def write_excel(rep_urls, rep_hashes, rep_ips, output="zz.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reputation"

    ws.append(["Type", "Indicator", "VirusTotal"])

    for url, vt in rep_urls:
        ws.append(["URL", url, vt])

    for h, vt in rep_hashes:
        ws.append(["HASH", h, vt])

    for ip, vt in rep_ips:
        ws.append(["IP", ip, vt])

    # Formatting
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color="4B0082", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border

    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 3

    wb.save(output)

    # Auto-open
    try:
        if platform.system() == "Windows":
            os.startfile(output)
        elif platform.system() == "Darwin":
            subprocess.call(["open", output])
        else:
            subprocess.call(["xdg-open", output])
    except:
        pass

# ==================================================
# MAIN
# ==================================================

if __name__ == "__main__":
    msg_file = "sample.msg"
    headers_file = "raw_headers.txt"

    api = load_api_keys()
    vt_key = api["VIRUSTOTAL_API_KEY"]

    urls, hashes = extract_msg_indicators(msg_file)
    ips = extract_public_ips(headers_file)

    rep_urls = [(u, vt_url(vt_key, u)) for u in urls]
    rep_hashes = [(h, vt_hash(vt_key, h)) for h in hashes]
    rep_ips = [(ip, vt_ip(vt_key, ip)) for ip in ips]

    write_excel(rep_urls, rep_hashes, rep_ips, output="zz.xlsx")
