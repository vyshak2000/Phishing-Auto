from striprtf.striprtf import rtf_to_text
from email.parser import Parser
from email.policy import default
from pathlib import Path
import pandas as pd
from openpyxl.styles import PatternFill

# ---------------- CONFIG ----------------
INPUT_FILE = "raw_headers.rtf"
OUTPUT_FILE = "headers_analysis.xlsx"
# ----------------------------------------


def read_rtf_headers(file_path):
    rtf_content = Path(file_path).read_text(encoding="utf-8", errors="ignore")
    return rtf_to_text(rtf_content).strip()


def parse_headers(raw_headers):
    parser = Parser(policy=default)
    return parser.parsestr(raw_headers)


def headers_to_dataframe(msg):
    rows = []
    for key, value in msg.items():
        rows.append({
            "Header": key,
            "Value": value.replace("\n", " ").strip()
        })
    return pd.DataFrame(rows)


def received_to_dataframe(msg):
    received_headers = msg.get_all("Received", [])
    rows = []
    for i, hop in enumerate(received_headers, start=1):
        rows.append({
            "Hop": i,
            "Details": hop.replace("\n", " ").strip()
        })
    return pd.DataFrame(rows)


def auth_to_dataframe(msg):
    auth_header = msg.get("Authentication-Results")

    if not auth_header:
        return pd.DataFrame(columns=["Check", "Result"])

    rows = []
    for part in auth_header.split(";"):
        part = part.strip()
        if "=" in part:
            check, result = part.split("=", 1)
            rows.append({
                "Check": check.strip(),
                "Result": result.strip()
            })

    return pd.DataFrame(rows)


def apply_auth_coloring(writer):
    sheet = writer.book["Authentication"]

    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    orange = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    grey = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2):
        cell = row[0]
        if not cell.value:
            continue

        value = cell.value.lower()

        if "pass" in value:
            cell.fill = green
        elif "fail" in value:
            cell.fill = red
        elif any(x in value for x in ["softfail", "neutral", "temperror"]):
            cell.fill = orange
        elif "none" in value:
            cell.fill = grey


def export_to_excel(headers_df, received_df, auth_df):
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        headers_df.to_excel(writer, sheet_name="Headers", index=False)
        received_df.to_excel(writer, sheet_name="Received_Hops", index=False)
        auth_df.to_excel(writer, sheet_name="Authentication", index=False)

        apply_auth_coloring(writer)


# ---------------- MAIN ----------------
if __name__ == "__main__":
    raw_headers = read_rtf_headers(INPUT_FILE)
    msg = parse_headers(raw_headers)

    headers_df = headers_to_dataframe(msg)
    received_df = received_to_dataframe(msg)
    auth_df = auth_to_dataframe(msg)

    export_to_excel(headers_df, received_df, auth_df)

    print(f"✔ Excel exported successfully: {OUTPUT_FILE}")
